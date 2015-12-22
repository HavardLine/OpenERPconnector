import Rest
from os.path import expanduser
from openpyxl import load_workbook


con = Rest.Connection()

#1: asking for every task_id, 2: asking for every work-record
ids = con.search('project.task.work',12)
print ids
works = con.get('project.task.work', [ids])
print works[1]



basePath = expanduser("~")+'\\server\\LTS\\grunndata\\'
templatePath = basePath+'template.xlsx'
outPath = basePath+'timesheets2.xlsx'

wb = load_workbook(templatePath) 
shResult = wb.get_sheet_by_name('result')
shPayment = wb.get_sheet_by_name('payment')

shCmd = wb.get_sheet_by_name('cmd')

#Overskrifter raw data
shResult.cell(row = 1, column = 1).value = 'Dato'
shResult.cell(row = 1, column = 2).value = 'Ansatt'
shResult.cell(row = 1, column = 3).value = 'Timer'
shResult.cell(row = 1, column = 4).value = 'Prosjekt/oppgave'
shResult.cell(row = 1, column = 5).value = 'Kommentar/logg'
shResult.cell(row = 1, column = 6).value = u'Timesl\xf8nn'

#Variabler tilknyttet lonn og timeantall

lonn = 0
timerBard = 0
timerHavard = 0
timerBardWideWTAC = 0
timerHavardWideWTAC = 0
timerBardWideCDS = 0
timerHavardWideCDS = 0
timerBardDynatec065 = 0
timerHavardDynatec065 = 0
timerBardDynatec096 = 0
timerHavardDynatec096 = 0
timerBardSlattland31159 = 0
timerHavardSlattland31159 = 0
timerBardInterntid = 0
timerHavardInterntid = 0


#loop all work records

for index, item in enumerate(works):
  print item['user_id'][1]
  shResult.cell(row = index+2, column = 1).value=item['date']
  shResult.cell(row = index+2, column = 2).value=item['user_id'][1]
  shResult.cell(row = index+2, column = 3).value=item['hours']
  shResult.cell(row = index+2, column = 4).value=item['task_id'][1]
  shResult.cell(row = index+2, column = 5).value=item['display_name']

#Timeantall Havard      
  if item['user_id'][0] == 5:
    lonn = 300
    print item['task_id']
    print u'Timesl\xf8nn H\xe5vard:', lonn, '\n'
    timerHavard = timerHavard + item['hours']
    if item['task_id'][0] == 1:
      timerHavardWideCDS = timerHavardWideCDS + item ['hours']
    elif item['task_id'][0] == 5:
      timerHavardWideWTAC = timerHavardWideWTAC + item ['hours']
    elif item['task_id'][0] == 4:
      timerHavardDynatec065 = timerHavardDynatec065 + item ['hours']
    elif item['task_id'][0] == 3:
      timerHavardDynatec096 = timerHavardDynatec096 + item ['hours']
    elif item['task_id'][0] == 6:
      timerHavardInterntid = timerHavardInterntid + item ['hours']
    elif item['task_id'][0] == 2:
      timerHavardSlattland31159 = timerHavardSlattland31159 + item ['hours']
  
#Timeantall Bard  
  elif item['user_id'][0] == 6:
    lonn = 250
    print item['task_id']
    print u'Timesl\xf8nn B\xe5rd:', lonn, '\n'
    timerBard = timerBard + item['hours']
    if item['task_id'][0] == 1:
      timerBardWideCDS = timerBardWideCDS + item ['hours']
    elif item['task_id'][0] == 5:
      timerBardWideWTAC = timerBardWideWTAC + item ['hours']
    elif item['task_id'][0] == 4:
      timerBardDynatec065 = timerBardDynatec065 + item ['hours']
    elif item['task_id'][0] == 3:
      timerBardDynatec096 = timerBardDynatec096 + item ['hours']
    elif item['task_id'][0] == 6:
      timerBardInterntid = timerBardInterntid + item ['hours']
    elif item['task_id'][0] == 2:
      timerBardSlattland31159 = timerBardSlattland31159 + item ['hours']
   
  else:
    print 'Fant ingen ansatt', '\n'
    
  shResult.cell(row = index+2, column = 6).value=lonn
print u'Timer H\xe5vard:', timerHavard
print u'Timer B\xe5rd:', timerBard
print 'Timer Dynatec 096 Havard:', timerHavardDynatec096
print 'Timer Dynatec 096 Bard:', timerBardDynatec096
print u'Timer Sl\xe5ttland 31159 H\xe5vard :', timerHavardSlattland31159
print u'Timer Sl\xe5ttland 31159 B\xe5rd:', timerBardSlattland31159
print u'Timer interntid H\xe5vard:', timerHavardInterntid
print u'Timer interntid B\xe5rd:', timerBardInterntid

shPayment.cell(row = 1, column = 1).value = 'Payment'
shPayment.cell(row = 2, column = 1).value = 'user_id'
shPayment.cell(row = 2, column = 2).value = 'prosjekt'
shPayment.cell(row = 2, column = 3).value = 'task_id'
shPayment.cell(row = 2, column = 4).value = 'hours'
shPayment.cell(row = 2, column = 5).value = 'total hours'

shPayment.cell(row = 3, column = 1).value = u'H\xe5vard B Line'
shPayment.cell(row = 3, column = 2).value = 'Wide'
shPayment.cell(row = 3, column = 3).value = 'Water to air cooler'
shPayment.cell(row = 3, column = 4).value = timerHavardWideWTAC
shPayment.cell(row = 3, column = 5).value = timerHavard
shPayment.cell(row = 4, column = 3).value = 'Collection duct scale'
shPayment.cell(row = 4, column = 4).value = timerHavardWideCDS
shPayment.cell(row = 5, column = 2).value = 'Dynatec'
shPayment.cell(row = 5, column = 3).value = '2015-065'
shPayment.cell(row = 5, column = 4).value = timerHavardDynatec065
shPayment.cell(row = 6, column = 3).value = '2014-096'
shPayment.cell(row = 6, column = 4).value = timerHavardDynatec096
shPayment.cell(row = 7, column = 2).value = U'Sl\xe5ttland'
shPayment.cell(row = 7, column = 3).value = '31159'
shPayment.cell(row = 7, column = 4).value = timerHavardSlattland31159
shPayment.cell(row = 8, column = 2).value = 'Interntid'
shPayment.cell(row = 8, column = 4).value = timerHavardInterntid

shPayment.cell(row = 13, column = 1).value = u'B\xe5rd Kopperud'
shPayment.cell(row = 13, column = 2).value = 'Wide'
shPayment.cell(row = 13, column = 3).value = 'Water to air cooler'
shPayment.cell(row = 13, column = 4).value = timerBardWideWTAC
shPayment.cell(row = 13, column = 5).value = timerBard
shPayment.cell(row = 14, column = 3).value = 'Collection duct scale'
shPayment.cell(row = 14, column = 4).value = timerBardWideCDS
shPayment.cell(row = 15, column = 2).value = 'Dynatec'
shPayment.cell(row = 15, column = 3).value = '2015-065'
shPayment.cell(row = 15, column = 4).value = timerBardDynatec065
shPayment.cell(row = 16, column = 3).value = '2014-096'
shPayment.cell(row = 16, column = 4).value = timerBardDynatec096
shPayment.cell(row = 17, column = 2).value = U'Sl\xe5ttland'
shPayment.cell(row = 17, column = 3).value = '31159'
shPayment.cell(row = 17, column = 4).value = timerBardSlattland31159
shPayment.cell(row = 18, column = 2).value = 'Interntid'
shPayment.cell(row = 18, column = 4).value = timerBardInterntid

wb.save(outPath)

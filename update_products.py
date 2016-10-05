import odoo, os, sys
from openpyxl import load_workbook

var_list_name = 'EUC'
path = os.environ['HOME']+'/'+var_list_name+'.xlsx'
print 'Reading from path: ' + path

wb = load_workbook(path, data_only=True)
ws = wb['Priser']

  #Setting up variables for loop
products = []
new_products = []
r=2
updates=0

#Reading spreadsheet
while ws.cell(row=r, column=5).value != None:
	products.append({
		'id': 'EUC.' + ws.cell(row=r, column=5).value.encode('ascii','ignore'),
		'name': ws.cell(row=r, column=7).value.encode('ascii','ignore') + ' ' + str(ws.cell(row=r, column=6).value),
		'standard_price': ws.cell(row=r, column=10).value,
		'list_price': ws.cell(row=r, column=14).value,
		'fields': {
		'standard_price': ws.cell(row=r, column=10).value,
		'list_price': ws.cell(row=r, column=14).value
		}
		})
	r+=1

print str(len(products)) + ' product lines found'

#Reading data from ODOO, updates price, separates new devices
con = odoo.Connection()

template = con.execute('product.template', 'search_read', [[['default_code', '=', 'LTS.part']]])
if len(template)==1:
	template = template[0]
	print 'Part template found'
else:
	print >> sys.stderr, "Failed to import part template"
	sys.exit(1)

for product in products:
	my_id = con.execute("product.template", "search", [[['default_code', '=', product['id']]]], {'limit': 3})
	if len(my_id) == 1:
		print str(product['id']) + ' ' + str(product['standard_price'])
		print my_id
		print con.execute("product.template","write",[my_id, {'standard_price':product['standard_price'], 'list_price':product['list_price']}])
		updates+=1
	elif len(my_id) > 1:
		print 'Product ' + str(product['id']) + ' has multiple variants and will be ignored'
	else:
		print 'Product ' + str(product['id']) + ' do not currently exist in ODOO'
		new_products.append({
			'active': template['active'],
			'default_code': product['id'], #input field
			'name': product['name'], #input field
			'standard_price': product['standard_price'], #input field
			'list_price': product['list_price'], #input field
			'mes_type': template['mes_type'],
			'uom_id': template['uom_id'][0],
			'uom_po_id': template['uom_po_id'][0],
			'type': template['type'],
			'cost_method': template['cost_method'],
			'categ_id': template['categ_id'][0]
			#'seller_ids':''
		})

print str(updates) + ' products updated'
print str(len(new_products)) + ' new products found'

#Verify and add new devices
from voluptuous import Schema, Coerce, ALLOW_EXTRA
schema = Schema({
                'default_code': str,
                'name': str,
                'list_price': Coerce(float),
                'standard_price': Coerce(float)
                }, extra=ALLOW_EXTRA)

for product_template in new_products:
	print 'Trying to add ' + product_template['default_code']
	schema(product_template)#fail if data is no good
	template_id = con.setProductTemplate([product_template])
	print 'template id = ' + str(template_id)
	print 'OK!!'

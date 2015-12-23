import Rest
from os.path import expanduser
from openpyxl import load_workbook

basePath = expanduser("~")+'\\server\\LTS\grunndata\\'
templatePath = basePath+'templateWorkReport.xlsx'
outPath = basePath+'workReport.xlsx'

wb = load_workbook(templatePath)
shReport = wb.get_sheet_by_name('report')

con = Rest.Connection(db='LTS', uri='http://172.16.1.146:8070')

#1: asking for every task_id, 2: asking for every work-record
ids = con.searchDate('project.task.work')
print ids
works = con.get('project.task.work', [ids])
print works[0]

#Sortere task_id foer listen skrives??

testarray = [5,8,6,14,2,9,10,11,11]
array = []

for index, item in enumerate(works):
  print item['display_name']
  print item['task_id']
  array.append(item['task_id'][0])
  
array.sort()
print array

#Ny loekke hvor prosjektene er sortert og kan skrives til excel
for index, item in enumerate(array):
  #print item['display_name']
  #print item['task_id']
  
  if item['task_id'][0] == 1:
    shReport.cell(row = 1, column = 1).value='Wide CDS'
    shReport.cell(row = index+1, column = 2).value=item['display_name']
    
  elif item['task_id'][0] == 5:
    shReport.cell(row = 10, column = 1).value='Wide WTAC'
    shReport.cell(row = index+1, column = 2).value=item['display_name']

wb.save(outPath)
